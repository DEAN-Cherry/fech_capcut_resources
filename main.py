import sys
import re
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def remove_prefix(id_string):
    return re.sub("table-item-", "", id_string)


def get_file_extension(filename):
    match = re.search(r'\.\w+$', filename)
    if match:
        return match.group()
    else:
        return None


def parse_html():
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = 'index.html'

    with open(file_path, encoding='utf-8') as file:
        _soup = BeautifulSoup(file, 'lxml')
        return _soup


def extract_data(_soup):
    _data = []
    ids = _soup.select('.lv-table-cell-wrap-value')
    for e in ids:
        if e.contents:
            div = e.contents[0]
            if div.name == 'div':
                item_id = remove_prefix(div.get('id'))
                des = div.select('.lv-typography')
                description = des[0].text if des else ''
                _data.append([item_id, description])

    return _data


def filter_data(_data):
    _filtered = []
    for i in _data:
        if not get_file_extension(i[1]):
            _filtered.append(i)
    return _filtered


def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    # for column_cells in ws.columns:
    #     new_column_length = max(len(str(cell.value)) for cell in column_cells)
    # new_column_letter = (get_column_letter(column_cells[0].column))
    # if new_column_length > 0:
    #     ws.column_dimensions[new_column_letter].width = new_column_length * 1.23
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column  # Get the column name
        for cell in column_cells:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


def dump_excel(_data, file_path='output.xlsx', cnt=0):
    try:
        if cnt > 0:
            file_path = file_path.replace('.xlsx', f'_{cnt}.xlsx')
        timestamp = pd.Timestamp.now().strftime('%Y-%m-%d %H-%M-%S')
        file_path = file_path.replace('.xlsx', f'_{timestamp}.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df = pd.DataFrame(_data, columns=['ID', 'Description'])
            df.to_excel(writer, sheet_name='All Data', index=False)

            invalid_df = pd.DataFrame(filter_data(_data), columns=['ID', 'Description'])
            invalid_df.to_excel(writer, sheet_name='Invalid Data', index=False)
            # 获取 "All Data" 工作表
            all_data_sheet = writer.book['All Data']

            # 获取 "Invalid Data" 工作表中的数据
            invalid_data = writer.book['Invalid Data'].values

            # 遍历 "Invalid Data" 中的数据
            for row in invalid_data:
                # 在 "All Data" 工作表中查找与 "Invalid Data" 中的 ID 匹配的行
                for cell in all_data_sheet['A'][1:]:
                    if cell.value == row[0]:
                        # 将匹配行的单元格填充荧光颜色
                        cell.fill = PatternFill('solid', fgColor='ffff00')
                        all_data_sheet[f'B{cell.row}'].fill = PatternFill('solid', fgColor='ffff00')

            columns_best_fit(all_data_sheet)
            columns_best_fit(writer.book['Invalid Data'])
    except PermissionError:
        cnt += 1
        dump_excel(_data, file_path, cnt)


if __name__ == '__main__':
    soup = parse_html()
    data = extract_data(soup)
    dump_excel(data)
    print('Done')
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
