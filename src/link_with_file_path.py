import os
import re
import sys
import pandas as pd
import utils.base_utils as utils
from tqdm import tqdm
from openpyxl.styles import PatternFill
from pathlib import Path


class FileParser:
    def __init__(self, html_path: str, local_resource_path: str, output_file_path: str = None, workspace_id: str ='7293489460916830210'):
        self.html_path = html_path
        self.local_resource_path = local_resource_path
        self.output_file_path = output_file_path
        self.workspace_id = workspace_id

        self.soup = utils.parse_html(html_path)
        self.data = utils.extract_data_2(self.soup)

        self.df_raw = pd.DataFrame(self.data, columns=['ID', 'Description'])
        self.df_raw.loc[:, 'URL'] = self.df_raw['ID'].apply(utils.generate_url, _workspace_id=workspace_id)

        self.df_invalid = pd.DataFrame(utils.filter_data(self.data), columns=['ID', 'Description'])
        self.df_name_url = self.df_raw[['Description', 'URL']]
        self.df_link = pd.DataFrame(columns=['File Path', 'File Name', 'URL'])

    def parse_file_number(self, file_name):
        match = re.search(r'_(\d+)\.\w+$', file_name)
        if match:
            return int(match.group(1))
        else:
            # TODO 这里有可能会出现不符合规则的文件名重复的情况，先不处理了
            # if (self.df_invalid['Description'] == file_name).any():
            #     return 0
            # else:
            return 0

    def link_local_file_with_url(self):
        for dir_path, dir_names, file_names in tqdm(os.walk(self.local_resource_path)):
            base_path = os.path.basename(dir_path)
            self.link_url_to_file_path_by_name(file_names, base_path)

        self.df_link.loc[:, 'file_number'] = self.df_link['File Name'].apply(self.parse_file_number)
        self.df_link = self.df_link.groupby('File Path').apply(lambda x: x.sort_values('file_number'), include_groups=False)
        self.df_link.reset_index(inplace=True)
        try:
            self.df_link = self.df_link.drop(columns=['file_number', 'level_1'])
        except KeyError:
            pass
            print("There in no relation data between file and URL.")

    def link_url_to_file_path_by_name(self, file_names, file_path):
        for file_name in file_names:
            _index = self.df_name_url[self.df_name_url['Description'].apply(lambda x: x.startswith(file_name))].index
            if not _index.empty:
                self.df_link.loc[_index[0]] = [file_path, file_name, self.df_name_url.loc[_index[0]]['URL']]

    def generate_file_path(self, cnt):
        if self.output_file_path and len(self.output_file_path) > 0:
            user_path = Path(self.output_file_path)
            if not user_path.exists():
                os.makedirs(user_path)
            file_path = user_path / Path('output.xlsx')
        else:
            file_path = 'output.xlsx'

        if cnt > 0:
            file_path = str(file_path).replace('.xlsx', f'_{cnt}.xlsx')
        timestamp = pd.Timestamp.now().strftime('%Y-%m-%d %H-%M-%S')

        file_path = str(file_path).replace('.xlsx', f'_{timestamp}.xlsx')

        return file_path

    def dump_excel(self, retry=0):
        try:
            excel_file_path = self.generate_file_path(retry)
            print(f'Excel file path: {Path(os.getcwd()) / Path(excel_file_path)}')
            with pd.ExcelWriter(f'{excel_file_path}', engine='openpyxl') as writer:
                self.df_link.to_excel(writer, sheet_name='All Data', index=False)
                self.df_invalid.to_excel(writer, sheet_name='Invalid Data', index=False)

                all_data_sheet = writer.book['All Data']
                invalid_data_sheet = writer.book['Invalid Data']
                invalid_data = invalid_data_sheet.values

                for row in invalid_data:
                    for cell in all_data_sheet['A'][1:]:
                        if cell.value == row[0]:
                            cell.fill = PatternFill('solid', fgColor='ffff00')
                            all_data_sheet[f'B{cell.row}'].fill = PatternFill('solid', fgColor='ffff00')

                utils.columns_best_fit(all_data_sheet)
                utils.columns_best_fit(writer.book['Invalid Data'])
        except PermissionError:
            retry += 1
            self.dump_excel(retry)

    def run(self):
        self.link_local_file_with_url()
        self.dump_excel()
