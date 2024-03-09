import sys
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from utils import base_utils as utils


def parse_html():
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = 'index.html'

    with open(file_path, encoding='utf-8') as file:
        _soup = BeautifulSoup(file, 'lxml')
        return _soup


def dump_excel(_data, file_path='output.xlsx', cnt=0):
    try:
        if cnt > 0:
            file_path = file_path.replace('.xlsx', f'_{cnt}.xlsx')
        timestamp = pd.Timestamp.now().strftime('%Y-%m-%d %H-%M-%S')
        file_path = file_path.replace('.xlsx', f'_{timestamp}.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df = pd.DataFrame(_data, columns=['ID', 'Description'])
            df.to_excel(writer, sheet_name='All Data', index=False)

            invalid_df = pd.DataFrame(utils.filter_data(_data), columns=['ID', 'Description'])
            invalid_df.to_excel(writer, sheet_name='Invalid Data', index=False)
            # 获取 "All Data" 工作表
            all_data_sheet = writer.book['All Data']

            # 获取 "Invalid Data" 工作表中的数据
            invalid_data = writer.book['Invalid Data'].values

            # 遍历 "Invalid Data" 中的数据
            for row in invalid_data:
                # 在 "All Data" 工作表中查找与 "Invalid Data" 中的 ID 匹配的行
                for cell in all_data_sheet['A'][1:]:
                    if cell.value == row[0]:
                        # 将匹配行的单元格填充荧光颜色
                        cell.fill = PatternFill('solid', fgColor='ffff00')
                        all_data_sheet[f'B{cell.row}'].fill = PatternFill('solid', fgColor='ffff00')

            utils.columns_best_fit(all_data_sheet)
            utils.columns_best_fit(writer.book['Invalid Data'])
    except PermissionError:
        cnt += 1
        dump_excel(_data, file_path, cnt)
