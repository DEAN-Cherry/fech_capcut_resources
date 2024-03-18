import openpyxl
import re
import bs4

from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


def remove_prefix(id_string):
    return re.sub("table-item-", "", id_string)


def parse_html(html_path):
    with open(html_path, encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'lxml')
        return soup


def get_file_extension(filename):
    match = re.search(r'\.\w+$', filename)
    if match:
        return match.group()
    else:
        return None


def generate_url(_item_id):
    # https://www.capcut.com/view/7342830990957478402?workspaceId=7293489460916830210&from=workspace
    return f'https://www.capcut.com/view/{_item_id}?workspaceId=7293489460916830210&from=workspace'


def extract_data_2(_soup):
    _data = []
    dataDiv = _soup.select('div[class*="DataViewGroupBody"]')
    for div in dataDiv:
        if div.children:
            for child in div.children:
                if isinstance(child, bs4.element.Tag):
                    # 获取 data-selectable-item-id 的值
                    item_id = child.get('data-selectable-item-id')
                    des = child.select('.lv-typography')
                    description = des[0].text if des else ''
                    if item_id is not None:  # 如果 data-selectable-item-id 的值存在
                        print(item_id)
                        print(description)
                        _data.append([item_id, description, generate_url(item_id)])
    return _data


def extract_data(_soup):
    _data = []
    ids = _soup.select('.lv-table-cell-wrap-value')
    for e in ids:
        if e.contents:
            div = e.contents[0]
            if div.name == 'div':
                item_id = remove_prefix(div.get('id'))
                des = div.select('.lv-typography')
                description = des[0].text if des else ''
                _data.append([item_id, description])

    return _data


def filter_data(_data):
    _filtered = []
    for i in _data:
        if not get_file_extension(i[1]):
            _filtered.append(i)
    return _filtered


def columns_best_fit(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    # for column_cells in ws.columns:
    #     new_column_length = max(len(str(cell.value)) for cell in column_cells)
    # new_column_letter = (get_column_letter(column_cells[0].column))
    # if new_column_length > 0:
    #     ws.column_dimensions[new_column_letter].width = new_column_length * 1.23
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column  # Get the column name
        for cell in column_cells:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (TypeError, AttributeError):
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
